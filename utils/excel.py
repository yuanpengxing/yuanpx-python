# -*- coding: UTF-8 -*-
# author: yuanpx

import warnings

import openpyxl
import xlrd
from openpyxl.chart import BarChart, Reference
from openpyxl_image_loader import SheetImageLoader

from utils.regex import RegexUtil


class OpenpylxUtil:
    def __init__(self, excelpath):
        warnings.filterwarnings('ignore')
        self.wb = openpyxl.load_workbook(excelpath)

    def set_sheet(self, sheet=0):
        self.st = self.wb.worksheets[sheet]

    def row_values(self, row):  # row=0表示第1行
        rows = self.st.rows
        data = []
        for i in range(len(rows)):
            if i == row:
                for cell in rows[i]:
                    data.append(cell.value)
        return data

    def col_values(self, col):  # col=0表示第1列
        cols = self.st.columns
        data = []
        for i in range(len(cols)):
            if i == col:
                for cell in cols[i]:
                    data.append(cell.value)
        return data

    def col_fg_colors(self, col):
        cols = self.st.columns
        data = []
        for i in range(len(cols)):
            if i == col:
                for cell in cols[i]:
                    data.append(cell.fill.fgColor.rgb)
        return data

    def rows_values(self):
        datas = []
        for row in self.st.rows:
            data = []
            for cell in row:
                data.append(cell.value)
            datas.append(data)
        return datas

    def max_row(self):
        return self.st.max_row

    def max_col(self):
        return self.st.max_column

    def close(self):
        self.wb.close()

    def save_image(self, row, path):
        loader = SheetImageLoader(self.st)
        keys = loader._images.keys()
        images = []
        for name in keys:
            if name:
                r = RegexUtil.extract(name, '\d+')[0]
                if float(r) == row:
                    try:
                        loader.get(name).save(path)
                        images.append(path)
                    except Exception as e:
                        pass
        return images

    def get_chart(self, ws, rows):
        chart = BarChart()
        chart.type, chart.style, chart.title = 'col', 10, rows[0][1].replace('\n', '')
        chart.y_axis.title = rows[0][1]
        maxr, maxc = len(rows), len(rows[0])
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=maxr, max_col=maxc), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=maxr))
        return chart

    def save_chart(self, rows, filexlsx):
        wb = openpyxl.load_workbook(filexlsx)
        for i in range(len(rows)):
            ws = wb['Sheet'] if i == 0 else wb['ART']
            row = rows[i]
            for values in row:
                ws.append(values)
            chart = self.get_chart(ws, row)
            ws.add_chart(chart, 'A10')
        wb.save(filexlsx)


class XlrdUtil:
    def __init__(self, fileName):
        self.wb = xlrd.open_workbook(fileName)

    def set_sheet(self, sheet=0):
        self.st = self.wb.sheet_by_index(sheet)

    def col_values(self, col):
        return self.st.col_values(col)

    def row_values(self, row):
        return self.st.row_values(row)

    def close(self):
        pass
